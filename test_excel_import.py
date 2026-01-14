#!/usr/bin/env python3
"""
Excel読み取り機能のテストスクリプト
運転資金・回転期間・資金繰り前提の読み取りをテスト
"""
import sys
import os

# プロジェクトルートをPythonパスに追加
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from app.utils.excel_import_helpers import (
    read_working_capital_assumptions,
    read_debt_repayment_assumptions
)


def test_read_working_capital_assumptions():
    """運転資金前提の読み取りテスト"""
    print("=" * 80)
    print("運転資金前提 読み取りテスト")
    print("=" * 80)
    
    excel_path = '/home/ubuntu/upload/簡易新エンジン3財務管理シミュレーション（個別計画あり）データあり20230301）_5.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"❌ Excelファイルが見つかりません: {excel_path}")
        return
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # 3年度分のデータを読み取る
    for i in range(3):
        year_name = ['初年度', '2年度', '3年度'][i]
        print(f"\n【{year_name}】")
        
        try:
            data = read_working_capital_assumptions(wb, i)
            
            print(f"  現預金回転期間: {data['cash_turnover_period']}")
            print(f"  売掛債権回転期間: {data['receivables_turnover_period']}")
            print(f"  棚卸資産回転期間: {data['inventory_turnover_period']}")
            print(f"  買掛債務回転期間: {data['payables_turnover_period']}")
            print(f"  手許現預金増加額: {data['cash_increase']:,}")
            print(f"  売掛債権増加額: {data['receivables_increase']:,}")
            print(f"  棚卸資産増加額: {data['inventory_increase']:,}")
            print(f"  買掛債務増加額: {data['payables_increase']:,}")
            
            print(f"  ✅ {year_name}の読み取り成功")
        except Exception as e:
            print(f"  ❌ {year_name}の読み取りエラー: {e}")
    
    print("\n" + "=" * 80)
    print("✅ 運転資金前提の読み取りテスト完了")
    print("=" * 80)


def test_read_debt_repayment_assumptions():
    """返済スケジュール前提の読み取りテスト"""
    print("\n" + "=" * 80)
    print("返済スケジュール前提 読み取りテスト")
    print("=" * 80)
    
    excel_path = '/home/ubuntu/upload/簡易新エンジン3財務管理シミュレーション（個別計画あり）データあり20230301）_5.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"❌ Excelファイルが見つかりません: {excel_path}")
        return
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # 3年度分のデータを読み取る
    for i in range(3):
        year_name = ['初年度', '2年度', '3年度'][i]
        print(f"\n【{year_name}】")
        
        try:
            data = read_debt_repayment_assumptions(wb, i)
            
            print(f"  借入金期首残高: {data['beginning_balance']:,}")
            print(f"  借入金借入額: {data['borrowing_amount']:,}")
            print(f"  借入金元本返済額: {data['principal_repayment']:,}")
            print(f"  借入金期末残高: {data['ending_balance']:,}")
            print(f"  支払利息: {data['interest_payment']:,}")
            print(f"  平均金利: {data['average_interest_rate']}")
            
            print(f"  ✅ {year_name}の読み取り成功")
        except Exception as e:
            print(f"  ❌ {year_name}の読み取りエラー: {e}")
    
    print("\n" + "=" * 80)
    print("✅ 返済スケジュール前提の読み取りテスト完了")
    print("=" * 80)


def test_cell_positions():
    """セル位置の検証テスト"""
    print("\n" + "=" * 80)
    print("セル位置検証テスト")
    print("=" * 80)
    
    excel_path = '/home/ubuntu/upload/簡易新エンジン3財務管理シミュレーション（個別計画あり）データあり20230301）_5.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"❌ Excelファイルが見つかりません: {excel_path}")
        return
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    print("\n【⑥主要運転資金計画】")
    ws = wb['⑥主要運転資金計画']
    
    # 初年度（列9）のセルを確認
    print("  初年度（列9）:")
    print(f"    Row 6 (現預金回転期間): {ws.cell(6, 9).value}")
    print(f"    Row 11 (売掛債権回転期間): {ws.cell(11, 9).value}")
    print(f"    Row 16 (棚卸資産回転期間): {ws.cell(16, 9).value}")
    print(f"    Row 22 (買掛債務回転期間): {ws.cell(22, 9).value}")
    
    print("\n【⑩資金調達返済計画（1）】")
    ws = wb['⑩資金調達返済計画（1）']
    
    # 列11のセルを確認
    print("  列11:")
    print(f"    Row 14 (借入金期首残高): {ws.cell(14, 11).value}")
    print(f"    Row 15 (借入金借入額): {ws.cell(15, 11).value}")
    print(f"    Row 16 (借入金元本返済額): {ws.cell(16, 11).value}")
    print(f"    Row 17 (借入金期末残高): {ws.cell(17, 11).value}")
    print(f"    Row 18 (支払利息): {ws.cell(18, 11).value}")
    print(f"    Row 39 (平均金利): {ws.cell(39, 11).value}")
    
    print("\n" + "=" * 80)
    print("✅ セル位置検証テスト完了")
    print("=" * 80)


def test_api_documentation():
    """APIドキュメント表示"""
    print("\n" + "=" * 80)
    print("API ドキュメント")
    print("=" * 80)
    
    print("\n【Excel読み取りAPI】")
    print("POST /decision/excel-import/working-capital-and-debt")
    print("パラメータ:")
    print("  - file: Excelファイル（multipart/form-data）")
    print("  - company_id: 企業ID")
    print("  - fiscal_year_ids: 会計年度IDのリスト（カンマ区切り、例: 1,2,3）")
    
    print("\n【運転資金前提取得API】")
    print("GET /decision/excel-import/working-capital/<fiscal_year_id>")
    print("パラメータ:")
    print("  - fiscal_year_id: 会計年度ID")
    
    print("\n【返済スケジュール前提取得API】")
    print("GET /decision/excel-import/debt-repayment/<fiscal_year_id>")
    print("パラメータ:")
    print("  - fiscal_year_id: 会計年度ID")
    
    print("\n【レスポンス例（運転資金前提）】")
    print("""
{
  "id": 1,
  "company_id": 1,
  "fiscal_year_id": 1,
  "cash_turnover_period": 0.033,
  "receivables_turnover_period": 2.4,
  "inventory_turnover_period": 0.73,
  "payables_turnover_period": 1.33,
  "cash_increase": 1234567.0,
  "receivables_increase": 2345678.0,
  "inventory_increase": 3456789.0,
  "payables_increase": 4567890.0
}
    """)
    
    print("\n【レスポンス例（返済スケジュール前提）】")
    print("""
{
  "id": 1,
  "company_id": 1,
  "fiscal_year_id": 1,
  "beginning_balance": 50000000.0,
  "borrowing_amount": 10000000.0,
  "principal_repayment": 5000000.0,
  "ending_balance": 55000000.0,
  "interest_payment": 1000000.0,
  "average_interest_rate": 0.02
}
    """)
    
    print("\n【使用例（curl）】")
    print("""
# Excel読み取り
curl -X POST 'http://localhost:5000/decision/excel-import/working-capital-and-debt' \\
  -F 'file=@/path/to/excel.xlsx' \\
  -F 'company_id=1' \\
  -F 'fiscal_year_ids=1,2,3'

# 運転資金前提取得
curl 'http://localhost:5000/decision/excel-import/working-capital/1'

# 返済スケジュール前提取得
curl 'http://localhost:5000/decision/excel-import/debt-repayment/1'
    """)
    
    print("\n【データベーステーブル】")
    print("1. working_capital_assumptions")
    print("   - 運転資金前提を保存")
    print("   - 回転期間（現預金、売掛債権、棚卸資産、買掛債務）")
    print("   - 運転資金増減額（現預金、売掛債権、棚卸資産、買掛債務）")
    print()
    print("2. debt_repayment_assumptions")
    print("   - 返済スケジュール前提を保存")
    print("   - 借入金明細（期首残高、借入額、元本返済額、期末残高）")
    print("   - 支払利息、平均金利")
    
    print("\n【注意事項】")
    print("- シート名は固定（推測なし）")
    print("- セル位置は固定（推測なし）")
    print("- 既存フィールドは変更なし")
    print("- 3年度分のデータを一括取り込み")
    
    print("\n" + "=" * 80)
    print("✅ APIドキュメント表示完了")
    print("=" * 80)


if __name__ == "__main__":
    test_cell_positions()
    test_read_working_capital_assumptions()
    test_read_debt_repayment_assumptions()
    test_api_documentation()
    
    print("\n" + "=" * 80)
    print("✅ すべてのテストが完了しました！")
    print("=" * 80)
