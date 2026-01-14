"""
Excelから複数年度の財務データを読み取って最小二乗法分析に渡すモジュール
"""

from typing import List, Dict, Any
import openpyxl


def import_multi_year_financial_data(file_path: str) -> Dict[str, Any]:
    """
    Excelファイルから複数年度の財務データを読み取る
    
    Args:
        file_path: Excelファイルのパス
    
    Returns:
        複数年度の財務データを含む辞書
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # PLシートから売上高、原価、販管費などを読み取る
    pl_sheet_name = "①PL"
    if pl_sheet_name not in wb.sheetnames:
        raise ValueError(f"シート '{pl_sheet_name}' が見つかりません")
    
    pl_sheet = wb[pl_sheet_name]
    
    # 複数年度のデータを読み取る（3期分）
    fiscal_years = []
    
    for fiscal_year_index in range(3):
        # 列のオフセット（初年度=K列、2年度=L列、3年度=M列）
        col_offset = 11 + fiscal_year_index  # K=11, L=12, M=13
        
        # 年度ラベルを取得
        year_label_cell = pl_sheet.cell(row=3, column=col_offset)
        year_label = year_label_cell.value if year_label_cell.value else f"Year {fiscal_year_index + 1}"
        
        # 売上高（Row 5）
        sales = pl_sheet.cell(row=5, column=col_offset).value or 0
        
        # 売上原価（Row 6）
        cost_of_sales = pl_sheet.cell(row=6, column=col_offset).value or 0
        
        # 売上総利益（Row 7）
        gross_profit = pl_sheet.cell(row=7, column=col_offset).value or 0
        
        # 販売費及び一般管理費（Row 8）
        operating_expenses = pl_sheet.cell(row=8, column=col_offset).value or 0
        
        # 営業利益（Row 9）
        operating_income = pl_sheet.cell(row=9, column=col_offset).value or 0
        
        # 営業外収益（Row 10）
        non_operating_income = pl_sheet.cell(row=10, column=col_offset).value or 0
        
        # 営業外費用（Row 11）
        non_operating_expenses = pl_sheet.cell(row=11, column=col_offset).value or 0
        
        # 経常利益（Row 12）
        ordinary_income = pl_sheet.cell(row=12, column=col_offset).value or 0
        
        # 特別利益（Row 13）
        extraordinary_income = pl_sheet.cell(row=13, column=col_offset).value or 0
        
        # 特別損失（Row 14）
        extraordinary_losses = pl_sheet.cell(row=14, column=col_offset).value or 0
        
        # 税引前当期純利益（Row 15）
        income_before_tax = pl_sheet.cell(row=15, column=col_offset).value or 0
        
        # 法人税等（Row 16）
        income_tax = pl_sheet.cell(row=16, column=col_offset).value or 0
        
        # 当期純利益（Row 17）
        net_income = pl_sheet.cell(row=17, column=col_offset).value or 0
        
        fiscal_years.append({
            'year': fiscal_year_index + 1,
            'year_label': year_label,
            'sales': float(sales),
            'cost_of_sales': float(cost_of_sales),
            'gross_profit': float(gross_profit),
            'operating_expenses': float(operating_expenses),
            'operating_income': float(operating_income),
            'non_operating_income': float(non_operating_income),
            'non_operating_expenses': float(non_operating_expenses),
            'ordinary_income': float(ordinary_income),
            'extraordinary_income': float(extraordinary_income),
            'extraordinary_losses': float(extraordinary_losses),
            'income_before_tax': float(income_before_tax),
            'income_tax': float(income_tax),
            'net_income': float(net_income)
        })
    
    wb.close()
    
    return {
        'fiscal_years': fiscal_years,
        'data_count': len(fiscal_years)
    }


def prepare_data_for_least_squares(financial_data: Dict[str, Any], metric: str) -> List[Dict[str, Any]]:
    """
    最小二乗法分析用にデータを整形
    
    Args:
        financial_data: import_multi_year_financial_data()の戻り値
        metric: 分析する指標名（例: 'cost_of_sales', 'operating_expenses'）
    
    Returns:
        最小二乗法分析用のデータリスト
    """
    result = []
    
    for year_data in financial_data['fiscal_years']:
        result.append({
            'year': year_data['year'],
            'sales': year_data['sales'],
            metric: year_data.get(metric, 0)
        })
    
    return result
